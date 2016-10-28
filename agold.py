# -*- coding: cp1251 -*-
import os, re
import openpyxl as op
import xlrd

from common_functions import *
                    
def getInfo(cell_index, ws):
    d = {u'Номер документа':ws.cell_value(1, 16),
         u'Дата документа':ws.cell_value(1, 17)}
    return d
    


def getData(ws, info, strt_row = 0):
    all_data = []
    finrow = 1
    while finrow < ws.nrows:
        artic = ws.cell_value(finrow, 2)
        name = ws.cell_value(finrow, 3)
        weight = ws.cell_value(finrow, 7)
        size = ws.cell_value(finrow, 5)
        price = ws.cell_value(finrow, 9)
        amount = ws.cell_value(finrow, 6)
        strich_code = repr(ws.cell_value(finrow, 1))

        strich_code = re.sub(u'u', u'', strich_code, flags=re.U)
        strich_code = re.sub(u' ', u'', strich_code, flags=re.U)
        strich_code = re.sub(u"'", u'', strich_code, flags=re.U)
        strich_code = re.sub(u"\.0", u'', strich_code, flags=re.U)

        name_cell = '%s/%s/%s/%s/%s'%(artic, #articul
                                        name, #type
                                        size, #size
                                        weight, #weight
                                        strich_code) #strich_code

        result_items = [artic,
                        name_cell,
                        weight,
                        size,
                        strich_code,
                        price,
                        amount]
        
        all_data.append(result_items)
##        print(all_data)
        finrow += 1
        
    return all_data



def makeResultWb(wbname, info, data, from_):
    wb = op.Workbook()
    ws = wb.active
    ws['A1'].value, ws['B1'].value = u'Номер документа', info[u'Номер документа']
    ws['A2'].value, ws['B2'].value = u'Дата документа', info[u'Дата документа']
    ws['A3'].value, ws['B3'].value = u'От кого', from_
    
    row = 4
    ws['A%s'%row] = u'Артикул поставщика'
    ws['B%s'%row] = u'Наименование'
    ws['C%s'%row] = u'Вес'
    ws['D%s'%row] = u'Размер'
    ws['E%s'%row] = u'Штрих-код'
    ws['F%s'%row] = u'Цена, руб. коп.'
    ws['G%s'%row] = u'Кол-во Изд-й'
    row += 1
    cols = 'ABCDEFG'
    for d in data:
        ws['A%s'%row] = d[0]
        ws['B%s'%row] = d[1]
        ws['C%s'%row] = d[2]
        ws['D%s'%row] = d[3]
        ws['E%s'%row] = d[4]
        ws['F%s'%row] = d[5]
        ws['G%s'%row] = d[6]
        row += 1
    wb.save(os.getcwd()+'\\results\\'+resultName(wbname)+'.xlsx')
    return True
        
        


def main(t, from_):
    data = 0
    findata = []
    row = 0
    ws = getWsData(t)
    cell_index = 0
    info = getInfo(cell_index, ws)
    info[u'От кого'] = from_
##    for i in info:
##        print i, info[i]
    findata = getData(ws, info)
    makeResultWb(t, info, findata, from_)
    return findata

if __name__ == '__main__':
    main()

