# -*- coding: cp1251 -*-
import os, re
import openpyxl as op
import xlrd

from common_functions import *

def getInfo(cell_index, ws):
    ddoc, ndoc = None, None
    d = {u'Номер документа':None, u'Дата документа':None}
    r = cell_index['row']
    c = cell_index['col']
    
    cnum = c +1 
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    d[u'Номер документа'] = ws.cell_value(r,cnum)
    cnum += 1
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    d[u'Дата документа'] = ws.cell_value(r,cnum)
    return d
    
def getSize(line):
    m = re.search(u'([A-Z]\(([^\)]+)\))', line)
    if m is not None:
        return m.group(2)
    else:
        return ''

def getData(ws, info, strt_row = 0):
    all_data = []
    name_index = findInWs(u'наименование, характеристика', ws, strt_row=strt_row)
    name_col_index = name_index['col']
##    print name_col_index
    row = name_index['row'] + 2
       
    try:
        strich_code_col_index = findInWs(u'штрих', ws)['col']
    except:
        strich_code_cell = findInWs(u'код$', ws, match=True)
##        print('strich code: ', strich_code_cell)
        strich_code_col_index = strich_code_cell['col']
##    print('strich code col index %s'%strich_code_col_index)

    try:
        artic_col_index = findInWs(u'код$', ws)['col']
    except:
        artic_col_index = strich_code_col_index
 
    weight_col_index = findInWs(u'Масса|Вес', ws)['col']
    amount_col_index = findInWs(u'Количество', ws)['col']
    price_col_index = findInWs(u'Сумма с', ws)['col']

    while ws.cell_value(row, name_col_index) != u'':
        name = ws.cell_value(row, name_col_index)
        artic = repr(ws.cell_value(row, artic_col_index))
        weight = ws.cell_value(row, weight_col_index)
        size = getSize(name)
        try:
            strich_code = repr(ws.cell_value(row, strich_code_col_index))
        except Exception as e:
##            print('cant find strich code due to %s'%e)
            strich_code = u''
        price = ws.cell_value(row, price_col_index)
        amount = ws.cell_value(row, amount_col_index)

        artic = re.sub(u'\.0', u'', artic, flags=re.U)
        artic = re.sub(u'u', u'', artic, flags=re.U)
        artic = re.sub(u"'", u'', artic, flags=re.U)
        strich_code = re.sub(u'\.0', u'', strich_code, flags=re.U)
    
        name_cell = '%s/%s/%s/%s/%s'%(artic, #articul
                                        name, #type
                                        size, #size
                                        weight, #weight
                                        strich_code) #strich_code

        name_cell = re.sub(u'\.0', u'', name_cell, flags=re.U)

        result_items = [artic,
                        name_cell,
                        weight,
                        size,
                        strich_code,
                        price,
                        amount]
                        
        all_data.append(result_items)
        
        row += 1
    finrow = row
##    print 'finrow', finrow
##    for i in all_data:
##        print i
    return all_data, finrow



def makeResultWb(wbname, info, data, from_):
    wb = op.Workbook()
    ws = wb.active
    ws['A1'].value, ws['B1'].value = u'Номер документа', info[u'Номер документа']
    ws['A2'].value, ws['B2'].value = u'Дата документа', info[u'Дата документа']
    ws['A3'].value, ws['B3'].value = u'От кого', from_
    
    row = 4
    ws['A%s'%row] = u'Артикул поставщика'
    ws['B%s'%row] = u'Наименование'
    ws['C%s'%row] = u'Масса'
    ws['D%s'%row] = u'Размер'
    ws['E%s'%row] = u'Штрих-код'
    ws['F%s'%row] = u'Цена, руб. коп.'
    ws['G%s'%row] = u'Кол-во Изд-й'
    row += 1
    cols = 'ABCDEFG'
    for d in data:
        ws['A%s'%row] = d[0]
        ws['B%s'%row] = d[1]
        ws['C%s'%row] = d[2]
        ws['D%s'%row] = getSize(d[1])
        ws['E%s'%row] = d[4]
        ws['F%s'%row] = d[5]
        ws['G%s'%row] = d[6]
        row += 1
    wb.save(os.getcwd()+'\\results\\'+resultName(wbname)+'.xlsx')
    return True
        
        


def main(t, from_):
    tables = 0
    data = 0
    findata = []
    row = 0
    ws = getWsData(t)
    cell_index = findInWs(u'ТОВАРНАЯ НАКЛАДНАЯ', ws)
    info = getInfo(cell_index, ws)
    info[u'От кого'] = from_
##        for i in info:
##            print i, info[i]
    while data is not None:
        try:
            data_and_row = getData(ws, info, row)
            data, row = data_and_row[0], data_and_row[1]
            findata += data
            tables += 1
        except Exception as e:
##            print e
            data = None
##        print len(findata)
##        print 'tables extracted: %s'%tables
    makeResultWb(t, info, findata, from_)

if __name__ == '__main__':
    main()
