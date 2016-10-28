# -*- coding: cp1251 -*-
import os, re
import openpyxl as op
import xlrd

from common_functions import *
                    
def getInfo(cell_index, ws):
    ddoc, ndoc = None, None
    d = {u'Номер документа':None, u'Дата документа':None}
    r = cell_index['row']
    c = cell_index['col']
    
    cnum = c +1 
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    d[u'Номер документа'] = ws.cell_value(r,cnum)
    cnum += 1
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    d[u'Дата документа'] = ws.cell_value(r,cnum)
    return d
    
def getSize(line):
    m = re.search(u'Размер: ([0-9][0-9]?,?[0-9]?[0-9]?)', line)
    if m is not None:
        return m.group(1)
    else:
        return 'no size'

def getData(ws, info, strt_row = 0):
    all_data = []
    name_index = findInWs(u'наименование, характеристика', ws, strt_row=strt_row)
    name_col_index = name_index['col']
    row = name_index['row'] + 2
    strich_code_col_index = findInWs(u'код', ws)['col']
    weight_col_index = findInWs(u'нетто', ws)['col']
    amount_col_index = findInWs(u'мест,', ws)['col']
    price_col_index = findInWs(u'Сумма', ws)['col']

    while ws.cell_value(row, name_col_index) != u'':
        
        name = ws.cell_value(row, name_col_index)
        m = re.search(u'([0-9]+) ', name)
        if m is not None:
            artic = m.group(1)
        else:
            artic = 'no artic'
            
        weight = ws.cell_value(row, weight_col_index)
        size = getSize(name)
        try:
            strich_code = repr(ws.cell_value(row, strich_code_col_index))[:-2]
        except:
            strich_code = u'no strich code'
        price = ws.cell_value(row, price_col_index)
        amount = ws.cell_value(row, amount_col_index)
    
        name_cell = '%s/%s/%s/%s/%s'%(str(artic), #articul
                                        name, #type
                                        size, #size
                                        weight, #weight
                                        strich_code) #strich_code

        result_items = [artic,
                        name_cell,
                        weight,
                        size,
                        strich_code,
                        price,
                        amount]
                        
        all_data.append(result_items)
        
        row += 1
    finrow = row

    for i in range(len(all_data)):
        s = i
        try:
            while all_data[s][3] == 'no size':
                s += 1
            all_data[i][3] = all_data[s][3]
            i = s + 1
        except:
            pass

    newdata = []
    for i in all_data:
        if i[5] != u'':
            if i[3] == 'no size':
                i[3] = ''
            i[1] = re.sub(u'no size', i[3], i[1], flags=re.U)
            newdata.append(i)
        else:
            pass
    all_data = newdata
    return all_data, finrow



def makeResultWb(wbname, info, data, from_):
    wb = op.Workbook()
    ws = wb.active
    ws['A1'].value, ws['B1'].value = u'Номер документа', info[u'Номер документа']
    ws['A2'].value, ws['B2'].value = u'Дата документа', info[u'Дата документа']
    ws['A3'].value, ws['B3'].value = u'От кого', from_
    
    row = 4
    ws['A%s'%row] = u'Артикул поставщика'
    ws['B%s'%row] = u'Наименование'
    ws['C%s'%row] = u'Масса'
    ws['D%s'%row] = u'Размер'
    ws['E%s'%row] = u'Штрих-код'
    ws['F%s'%row] = u'Цена, руб. коп.'
    ws['G%s'%row] = u'Кол-во Изд-й'
    row += 1
    cols = 'ABCDEFG'
    for d in data:
        ws['A%s'%row] = d[0]
        ws['B%s'%row] = d[1]
        ws['C%s'%row] = d[2]
        ws['D%s'%row] = d[3]
        ws['E%s'%row] = d[4]
        ws['F%s'%row] = d[5]
        ws['G%s'%row] = d[6]
        row += 1
    wb.save(os.getcwd()+'\\results\\'+resultName(wbname)+'.xlsx')
    return True
    

def main(t, from_):
##    print(t)
    tables = 0
    data = 0
    findata = []
    row = 0
    ws = getWsData(t)
    cell_index = findInWs(u'ТОВАРНАЯ НАКЛАДНАЯ', ws)
    info = getInfo(cell_index, ws)
    info[u'От кого'] = from_
##        for i in info:
##            print i, info[i]
    while data is not None:
        try:
            data_and_row = getData(ws, info, row)
            data, row = data_and_row[0], data_and_row[1]
##            d_ = []
##            for i in data:
##                if i[0] != u'no artic':
##                    d_.append(i)
##            data = d_        
            findata += data
            tables += 1
        except Exception as e:
##            print e
            data = None
##        print len(findata)
##        print 'tables extracted: %s'%tables
    makeResultWb(t, info, findata, from_)

if __name__ == '__main__':
    main()
