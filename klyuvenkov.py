# -*- coding: cp1251 -*-
import os, re
import openpyxl as op
import xlrd

from common_functions import *
                    
def getInfo(cell_index, ws):
    ddoc, ndoc = None, None
    d = {u'Номер документа':None, u'Дата документа':None}
    r = cell_index['row']
    c = cell_index['col']
    
    cnum = c + 1 
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    d[u'Номер документа'] = ws.cell_value(r,cnum)
    cnum += 1
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    cnum += 1
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    d[u'Дата документа'] = ws.cell_value(r,cnum)
    return d
    
def getSize(line):
    m = re.search(u'разм: ([0-9][0-9]?,?[0-9]?[0-9]?)', line)
    if m is not None:
        return m.group(1)
    else:
        return ''

def getData(ws, info, strt_row = 0):
    all_data = []
    name_index = findInWs(u'Наименование товара', ws, strt_row=strt_row)
    name_col_index = name_index['col']
    artic_cell = findInWs(u'Код товара', ws, strt_row=strt_row)
    artic_col_index = artic_cell['col']
    amount_col_index = findInWs(u'объем', ws, strt_row=strt_row)['col']
    price_col_index = findInWs(u'Цена', ws, strt_row=strt_row)['col']

##    print('into the fire')
    row = artic_cell['row']+3
    tab_strt_col = artic_cell['col']
    
    while ws.cell_value(row, tab_strt_col) != u'':
        strich_code = repr(ws.cell_value(row, artic_col_index))[:-2]
        
        name = ws.cell_value(row, name_col_index)
        m = re.search(u'арт. ([^ ]+) ', name)
        if m is not None:
            artic = m.group(1)
        else:
            artic = 'no article'
        m = re.search(u'вес: ([0-9,]+)', name)
        if m is not None:
            weight = m.group(1)
            if weight[-1] not in u'0123456789':
                weight = weight[:-1]
        else:
            weight = u''
        
        size = getSize(name)
        price = ws.cell_value(row, price_col_index)
        amount = ws.cell_value(row, amount_col_index)
    
        name_cell = '%s/%s/%s/%s/%s'%(artic, #articul
                                        name, #type
                                        size, #size
                                        weight, #weight
                                        strich_code) #strich_code

        result_items = [artic,
                        name_cell,
                        weight,
                        size,
                        strich_code,
                        price,
                        amount]

        
        all_data.append(result_items)
        
        
        row += 1
##    print('outta fire')
    finrow = row
    
    
    return all_data, finrow



def makeResultWb(wbname, info, data, from_):
    wb = op.Workbook()
    ws = wb.active
    ws['A1'].value, ws['B1'].value = u'Номер документа', info[u'Номер документа']
    ws['A2'].value, ws['B2'].value = u'Дата документа', info[u'Дата документа']
    ws['A3'].value, ws['B3'].value = u'От кого', from_
    
    row = 4
    ws['A%s'%row] = u'Артикул поставщика'
    ws['B%s'%row] = u'Наименование'
    ws['C%s'%row] = u'Масса'
    ws['D%s'%row] = u'Размер'
    ws['E%s'%row] = u'Штрих-код'
    ws['F%s'%row] = u'Цена, руб. коп.'
    ws['G%s'%row] = u'Кол-во Изд-й'
    row += 1
    cols = 'ABCDEFG'
    for d in data:
        ws['A%s'%row] = d[0]
        ws['B%s'%row] = d[1]
        ws['C%s'%row] = d[2]
        ws['D%s'%row] = d[3]
        ws['E%s'%row] = d[4]
        ws['F%s'%row] = d[5]
        ws['G%s'%row] = d[6]
        row += 1
    wb.save(os.getcwd()+'\\results\\'+resultName(wbname)+'.xlsx')
    return True
        
        

def main(t, from_):
    if os.path.splitext(t)[-1] == '.dbf' or os.path.splitext(t)[-1] == '.DBF':

        def getItem(i, key):
            try:
##                print(key, i[key])
                return i[key]
            except Exception as e:
##                print('getItem: %s'%e)
                return u''
        
        ws = getWsData(t)
        info = {u'Номер документа':u'', u'Дата документа':u''}
        info[u'От кого'] = from_
        all_data = []
        
        
        for i in ws:
            artic = getItem(i, 'ARTICUL')
            name = getItem(i, 'FULLNAME')
            size = getItem(i, 'SIZE')
            strich_code = getItem(i, 'BARCODE')
            weight = getItem(i, 'WEIGHT')
            price = getItem(i, 'PRICE')
            amount = getItem(i, 'COUNT')
            
            
            name_cell = '%s/%s/%s/%s/%s'%(artic, #articul
                                        name, #type
                                        size, #size
                                        weight, #weight
                                        strich_code) #strich_code
            result_items = [artic,
                        name_cell,
                        weight,
                        size,
                        strich_code,
                        price,
                        amount]

        
            all_data.append(result_items)
        makeResultWb(t, info, all_data, from_)


    else:            
        tables = 0
        data = 0
        findata = []
        row = 0
        ws = getWsData(t)
        cell_index = findInWs(u'Счет-фактура', ws)
        info = getInfo(cell_index, ws)
        info[u'От кого'] = from_
    ##        for i in info:
    ##            print(i, info[i])
        while data is not None:
            try:
                data_and_row = getData(ws, info, strt_row=row)
                data, row = data_and_row[0], data_and_row[1]
    ##                print('ROW', row)
    ##                print(data)
                findata += data
                tables += 1
            except Exception as e:
    ##            print e
                data = None
    ##        print(len(findata))
    ##        print('tables extracted: %s'%tables)
        new_findata = []
        artics = []
        for i in findata:
            if i[0] not in artics:
                artics.append(i[0])
                new_findata.append(i)
            else:
                pass
        findata = new_findata
        makeResultWb(t, info, findata, from_)

if __name__ == '__main__':
    main()
