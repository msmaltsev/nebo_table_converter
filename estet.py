# -*- coding: cp1251 -*-
import xlrd
import xlwt
import os, re
import openpyxl as op
import codecs

from common_functions import *


def getInfo(ws):
    ndoc = ws.cell_value(3,1)
    ddoc = ws.cell_value(5,1)
    fdoc = ws.cell_value(1, 1)
    return {u'Номер документа':ndoc, u'Дата документа':ddoc, u'От кого':fdoc}

def findSum(row, col):
##    print ws.cell_value(row, col)
    try:
        if u'Итого' in cell.value:
            return True
        else:
            return False
    except:
        return False

def getStrichCode(ws, line):
    return ws.cell_value(line+1,22)

def getData(ws):
    all_data = []
    line = 6
    while findSum(line, 0) is not True:
        if ws.cell_value(line, 0) != u'':
            try:
                strich_code = getStrichCode(ws, line)
                name_cell = '%s/%s/%s/%s/%s/%s'%(ws.cell_value(line, 0), #articul
                                        ws.cell_value(line, 1), #type
                                        ws.cell_value(line, 4), #size
                                        ws.cell_value(line, 2), #probe
                                        ws.cell_value(line, 5), #weight
                                        strich_code)
                result_items = [ws.cell_value(line, 0),
                              name_cell,
                              ws.cell_value(line, 5),
                              ws.cell_value(line, 4),
                              strich_code,
                              ws.cell_value(line, 7),
                              ws.cell_value(line, 3)]
##                print result_items
                all_data.append(result_items)
            except:
                break
            
        line += 1
    return all_data

def makeResultWb(wbname, info, data, from_):
    wb = op.Workbook()
    ws = wb.active
    ws['A1'].value, ws['B1'].value = u'Номер документа', info[u'Номер документа']
    ws['A2'].value, ws['B2'].value = u'Дата документа', info[u'Дата документа']
    ws['A3'].value, ws['B3'].value = u'От кого', from_
    
    row = 4
    ws['A%s'%row] = u'Артикул поставщика'
    ws['B%s'%row] = u'Наименование'
    ws['C%s'%row] = u'Вес'
    ws['D%s'%row] = u'Размер'
    ws['E%s'%row] = u'Штрих-код'
    ws['F%s'%row] = u'Цена'
    ws['G%s'%row] = u'Кол-во Изд-й'
    row += 1
    cols = 'ABCDEFG'
    for d in data:
        ws['A%s'%row] = d[0]
        ws['B%s'%row] = d[1]
        ws['C%s'%row] = d[2]
        ws['D%s'%row] = d[3]
        ws['E%s'%row] = d[4]
        ws['F%s'%row] = d[5]
        ws['G%s'%row] = d[6]
        row += 1
    wb.save(os.getcwd()+'\\results\\'+resultName(wbname)+'.xlsx')
    return True


    

def main(wb, from_):
    ws = getWsData(wb)
    info = getInfo(ws)
    info[u'От кого'] = from_
    data = getData(ws)
##    for d in data:
##        print d
    makeResultWb(wb, info, data, from_)

if __name__ == '__main__':
    main()

