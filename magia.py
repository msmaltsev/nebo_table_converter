# -*- coding: cp1251 -*-
import os, re
import openpyxl as op
import xlrd

from common_functions import *
                    
def getInfo(cell_index, ws):
    ddoc, ndoc = None, None
    d = {u'Номер документа':None, u'Дата документа':None}
    r = cell_index['row']
    c = cell_index['col']
    
    cnum = c +1 
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    d[u'Номер документа'] = ws.cell_value(r,cnum)
    cnum += 1
    while ws.cell_value(r,cnum) is None or ws.cell_value(r,cnum) == u'':
        cnum += 1
    d[u'Дата документа'] = ws.cell_value(r,cnum)
    return d
    
def getSize(line):
    m = re.search(u'р. ([0-9][0-9]?,?[0-9]?[0-9]?)', line)
    if m is not None:
        return m.group(1)
    else:
        return ''

def getData(ws, info, strt_row = 0):
    all_data = []
    name_index = findInWs(u'наименование, характеристика', ws, strt_row=strt_row)
    name_col_index = name_index['col']
    row = name_index['row'] + 2
    artic_col_index = findInWs(u'код', ws)['col']
    weight_col_index = findInWs(u'Масса', ws)['col']
    amount_col_index = findInWs(u'Количество', ws)['col']
    price_col_index = findInWs(u'НДС', ws)['col']


    while ws.cell_value(row, name_col_index) != u'':
        artic = ws.cell_value(row, artic_col_index)
        name = ws.cell_value(row, name_col_index)
        weight = ws.cell_value(row, weight_col_index)
        size = getSize(name)
        price = ws.cell_value(row, price_col_index)
        amount = ws.cell_value(row, amount_col_index)
    
        name_cell = '%s/%s/%s/%s/%s'%(artic, #articul
                                        name, #type
                                        size, #size
                                        weight, #weight
                                        '') #strich_code
        m = re.search(u'([^\.]+)(\.[^/]+)(/.*)', name_cell)
        if m is not None:
            name_cell = m.group(1)+m.group(3)

        result_items = [artic,
                        name_cell,
                        weight,
                        size,
                        '',
                        price,
                        amount]
                        
        all_data.append(result_items)
        
        row += 1
    finrow = row
    return all_data, finrow



def makeResultWb(wbname, info, data, from_):
    wb = op.Workbook()
    ws = wb.active
    ws['A1'].value, ws['B1'].value = u'Номер документа', info[u'Номер документа']
    ws['A2'].value, ws['B2'].value = u'Дата документа', info[u'Дата документа']
    ws['A3'].value, ws['B3'].value = u'От кого', from_
    
    row = 4
    ws['A%s'%row] = u'Артикул поставщика'
    ws['B%s'%row] = u'Наименование'
    ws['C%s'%row] = u'Вес'
    ws['D%s'%row] = u'Размер'
    ws['E%s'%row] = u'Штрих-код'
    ws['F%s'%row] = u'Цена, руб. коп.'
    ws['G%s'%row] = u'Кол-во Изд-й'
    row += 1
    cols = 'ABCDEFG'
    for d in data:
        ws['A%s'%row] = d[0]
        ws['B%s'%row] = d[1]
        ws['C%s'%row] = d[2]
        ws['D%s'%row] = d[3]
        ws['E%s'%row] = d[4]
        ws['F%s'%row] = d[5]
        ws['G%s'%row] = d[6]
        row += 1
    wb.save(os.getcwd()+'\\results\\'+resultName(wbname)+'.xlsx')
    return True
        
        


def main(t, from_):
    
    data = 0
    findata = []
    row = 0
    ws = getWsData(t)
    m = re.search(u'Название', ws.cell_value(3,0))
    if m is not None:
        info = {u'Номер документа':ws.cell_value(0,0),
                u'Дата документа':ws.cell_value(1,0),
                u'От кого':ws.cell_value(2,0)}
        for row in range(4,ws.nrows):
            artic = ws.cell_value(row, 1)
            name = ws.cell_value(row, 0)
            weight = ws.cell_value(row, 3)
            size = ws.cell_value(row, 4)
            price = ws.cell_value(row, 5)
            amount = ws.cell_value(row, 6)
            strich_code = ws.cell_value(row, 2)
        
            name_cell = '%s/%s/%s/%s/%s'%(artic, #articul
                                            name, #type
                                            size, #size
                                            weight, #weight
                                            strich_code) #strich_code

            result_items = [artic,
                        name_cell,
                        weight,
                        size,
                        strich_code,
                        price,
                        amount]

            
                        
            findata.append(result_items)
        makeResultWb(t, info, findata, info[u'От кого'])
    else:
        cell_index = findInWs(u'ТОВАРНАЯ НАКЛАДНАЯ', ws)
        info = getInfo(cell_index, ws)
        info[u'От кого'] = from_
    ##        for i in info:
    ##            print i, info[i]
        while data is not None:
            try:
                data_and_row = getData(ws, info, row)
                data, row = data_and_row[0], data_and_row[1]
                findata += data
            except:
    ##            print e
                data = None
    ##        print len(findata)
        makeResultWb(t, info, findata, from_)
    return findata

if __name__ == '__main__':
    main()

